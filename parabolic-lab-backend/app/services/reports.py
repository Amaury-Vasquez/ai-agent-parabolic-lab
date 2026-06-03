"""
Servicio para generación de reportes en CSV y PDF.
Incluye desempeño de estudiantes, parámetros físicos y estadísticas de éxito.
"""

import csv
import io
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.alumno import Alumno
from app.models.alumno_en_salon import AlumnoEnSalon
from app.models.escenario import Escenario
from app.models.interaccion_escenario import InteraccionEscenario
from app.models.salon import Salon
from app.schemas.interaccion_escenario import ReporteInteraccionRead

logger = logging.getLogger(__name__)


class ReportService:
    """Servicio para generar reportes de desempeño académico."""

    @staticmethod
    async def generate_salon_csv_report(db: AsyncSession, salon_id: str) -> bytes:
        """
        Genera un reporte CSV consolidado del salón.

        Incluye:
        - Estadísticas del salón
        - Resumen por estudiante
        - Desempeño en escenarios
        """
        # Obtener salón con estudiantes
        result = await db.execute(
            select(Salon)
            .where(Salon.idsalon == salon_id)
            .options(
                selectinload(Salon.alumnos)
                .selectinload(AlumnoEnSalon.alumno)
                .selectinload(Alumno.usuario)
            )
        )
        salon = result.scalar_one_or_none()
        if not salon:
            raise ValueError(f"Salón {salon_id} no encontrado")

        # Crear buffer para CSV
        output = io.StringIO()
        writer = csv.writer(output)

        # Encabezado general
        writer.writerow(["REPORTE DE DESEMPEÑO - SALÓN"])
        writer.writerow([])
        writer.writerow(["Salón:", salon.nombresalon])
        writer.writerow(["Código de Acceso:", salon.codigoacceso])
        writer.writerow(["Fecha Reporte:", datetime.now().strftime("%Y-%m-%d %H:%M")])
        writer.writerow([])

        # Encabezado de tabla de estudiantes
        writer.writerow(
            [
                "Nombre",
                "Apellido Paterno",
                "Apellido Materno",
                "Matrícula",
                "Salón",
                "Intentos Totales",
                "Escenarios Completados",
                "Promedio Puntuación",
                "Tiempo Total (min)",
                "Tasa Éxito (%)",
            ]
        )

        # Procesar cada estudiante
        for alumno_en_salon in salon.alumnos:
            if not alumno_en_salon.activo:
                continue
            alumno = alumno_en_salon.alumno
            usuario = alumno.usuario

            # Obtener interacciones del estudiante en escenarios de este salón
            result = await db.execute(
                select(InteraccionEscenario).where(
                    InteraccionEscenario.idalumno == alumno.idalumno,
                    InteraccionEscenario.idescenario.in_(
                        select(Escenario.idescenario).where(Escenario.idsalon == salon.idsalon)
                    ),
                )
            )
            interacciones = result.scalars().all()

            # Calcular estadísticas usando intentosrealizados reales
            total_intentos = sum(int(i.intentosrealizados or 0) for i in interacciones)
            completadas = sum(1 for i in interacciones if i.completado)
            num_interacciones = len(interacciones)

            if interacciones:
                puntuacion_promedio = (
                    sum(float(i.puntuacion or 0) for i in interacciones) / num_interacciones
                )
                tiempo_total_seg = sum(int(i.tiempototal or 0) for i in interacciones)
                tasa_exito = (completadas / num_interacciones * 100) if num_interacciones > 0 else 0
            else:
                puntuacion_promedio = 0
                tiempo_total_seg = 0
                tasa_exito = 0

            writer.writerow(
                [
                    usuario.nombre or "Sin nombre",
                    usuario.apellidopaterno or "",
                    usuario.apellidomaterno or "",
                    alumno.matricula,
                    salon.nombresalon,
                    total_intentos,
                    completadas,
                    round(puntuacion_promedio, 2),
                    round(tiempo_total_seg / 60, 2),
                    round(tasa_exito, 2),
                ]
            )

        # Convertir a bytes con BOM for Excel compatibility
        bom = "\ufeff"
        return (bom + output.getvalue()).encode("utf-8")

    @staticmethod
    async def generate_student_csv_report(db: AsyncSession, alumno_id: str, salon_id: str) -> bytes:
        """Genera reporte CSV de un estudiante específico en un salón."""
        # Obtener alumno
        result = await db.execute(
            select(Alumno).where(Alumno.idalumno == alumno_id).options(selectinload(Alumno.usuario))
        )
        alumno = result.scalar_one_or_none()
        if not alumno:
            raise ValueError(f"Estudiante {alumno_id} no encontrado")

        usuario = alumno.usuario

        # Obtener interacciones del estudiante en el salón
        result = await db.execute(
            select(InteraccionEscenario)
            .where(
                InteraccionEscenario.idalumno == alumno_id,
                InteraccionEscenario.idescenario.in_(
                    select(Escenario.idescenario).where(Escenario.idsalon == salon_id)
                ),
            )
            .options(selectinload(InteraccionEscenario.escenario))
        )
        interacciones = result.scalars().all()

        # Crear buffer
        output = io.StringIO()
        writer = csv.writer(output)

        # Encabezado
        nombre_completo = " ".join(
            filter(None, [usuario.nombre, usuario.apellidopaterno, usuario.apellidomaterno])
        )
        writer.writerow(["EXPEDIENTE DE DESEMPEÑO - ESTUDIANTE"])
        writer.writerow([])
        writer.writerow(["Nombre:", nombre_completo or "Sin nombre"])
        writer.writerow(["Matrícula:", alumno.matricula])
        writer.writerow(["Email:", usuario.email or "N/A"])
        writer.writerow(["Fecha Reporte:", datetime.now().strftime("%Y-%m-%d %H:%M")])
        writer.writerow([])

        if not interacciones:
            writer.writerow(["Sin datos registrados"])
            writer.writerow([])
            writer.writerow(["Este estudiante aún no ha participado en escenarios."])
        else:
            # Tabla de intentos
            writer.writerow(
                [
                    "Escenario",
                    "Nivel Dificultad",
                    "Velocidad Inicial (m/s)",
                    "Ángulo (°)",
                    "Intentos",
                    "Completado",
                    "Puntuación",
                    "Tiempo Total (seg)",
                    "Fecha",
                ]
            )

            for interaccion in interacciones:
                escenario = interaccion.escenario
                datos = interaccion.datosinteraccion or {}

                v0 = datos.get("velocidadInicial", "N/A")
                angulo = datos.get("angulo", "N/A")

                writer.writerow(
                    [
                        escenario.nombre,
                        escenario.niveldificultad,
                        v0,
                        angulo,
                        interaccion.intentosrealizados,
                        "Sí" if interaccion.completado else "No",
                        round(float(interaccion.puntuacion) if interaccion.puntuacion else 0, 2),
                        int(interaccion.tiempototal) if interaccion.tiempototal else 0,
                        interaccion.fechafin.strftime("%Y-%m-%d %H:%M") if interaccion.fechafin else "Sin registro",
                    ]
                )

        # BOM for Excel compatibility
        bom = "\ufeff"
        return (bom + output.getvalue()).encode("utf-8")

    @staticmethod
    async def generate_salon_pdf_report(db: AsyncSession, salon_id: str) -> bytes:
        """
        Genera un reporte PDF profesional del salón.
        Incluye tablas, estadísticas y parámetros físicos.
        """
        # Obtener datos del salón
        result = await db.execute(
            select(Salon)
            .where(Salon.idsalon == salon_id)
            .options(
                selectinload(Salon.alumnos)
                .selectinload(AlumnoEnSalon.alumno)
                .selectinload(Alumno.usuario)
            )
        )
        salon = result.scalar_one_or_none()
        if not salon:
            raise ValueError(f"Salón {salon_id} no encontrado")

        # Crear documento PDF
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=letter,
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#1F4788"),
            spaceAfter=12,
            alignment=1,  # Centrado
        )
        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=14,
            textColor=colors.HexColor("#2C5AA0"),
            spaceAfter=10,
            spaceBefore=10,
        )

        # Contenido del documento
        elements = []

        # Título
        elements.append(Paragraph("PARABOLIC LAB", title_style))
        elements.append(Paragraph("Reporte de Desempeño - Salón", styles["Heading2"]))
        elements.append(Spacer(1, 0.2 * inch))

        # Información del salón
        salon_info = [
            ["Salón:", salon.nombresalon],
            ["Código de Acceso:", salon.codigoacceso],
            ["Total Estudiantes:", len(salon.alumnos)],
            ["Fecha Reporte:", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ]
        salon_table = Table(salon_info, colWidths=[2 * inch, 4 * inch])
        salon_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8F0F8")),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ]
            )
        )
        elements.append(salon_table)
        elements.append(Spacer(1, 0.3 * inch))

        # --- Collect per-student data ---
        student_rows = []  # list of dicts with computed stats
        try:
            for alumno_en_salon in salon.alumnos:
                alumno = alumno_en_salon.alumno
                usuario = alumno.usuario
                nombre_completo = " ".join(
                    filter(None, [usuario.nombre, usuario.apellidopaterno, usuario.apellidomaterno])
                )

                # Obtener interacciones
                result = await db.execute(
                    select(InteraccionEscenario).where(
                        InteraccionEscenario.idalumno == alumno.idalumno,
                        InteraccionEscenario.idescenario.in_(
                            select(Escenario.idescenario).where(Escenario.idsalon == salon.idsalon)
                        ),
                    )
                )
                interacciones = result.scalars().all()

                total_intentos = sum(int(i.intentosrealizados or 0) for i in interacciones)
                completadas = sum(1 for i in interacciones if i.completado)
                num_interacciones = len(interacciones)

                if interacciones:
                    promedio = (
                        sum(float(i.puntuacion or 0) for i in interacciones) / num_interacciones
                    )
                    mejor = max((float(i.puntuacion or 0) for i in interacciones), default=0)
                    tiempo_seg = sum(int(i.tiempototal or 0) for i in interacciones)
                    tasa = (completadas / num_interacciones * 100) if num_interacciones > 0 else 0
                else:
                    promedio = 0
                    mejor = 0
                    tiempo_seg = 0
                    tasa = 0

                student_rows.append(
                    {
                        "nombre": nombre_completo or "Sin nombre",
                        "matricula": alumno.matricula,
                        "promedio": float(promedio),
                        "mejor": float(mejor),
                        "completados": int(completadas),
                        "intentos": int(total_intentos),
                        "tiempo_min": round(tiempo_seg / 60, 2),
                        "tasa": round(float(tasa), 1),
                        "num_interacciones": int(num_interacciones),
                    }
                )

            # --- Resumen General de Desempeño ---
            has_data = any(s["num_interacciones"] > 0 for s in student_rows)

            if len(salon.alumnos) == 0:
                elements.append(Paragraph("Sin datos registrados - El salón no tiene estudiantes.", styles["Normal"]))
            elif not has_data:
                elements.append(
                    Paragraph(
                        "No hay datos de interacción registrados aún para ningún estudiante.",
                        styles["Normal"],
                    )
                )
            else:
                elements.append(Paragraph("Resumen General de Desempeño", heading_style))

                resumen_data = [
                    [
                        "Estudiante",
                        "Promedio",
                        "Mejor Puntaje",
                        "Completados",
                        "Intentos",
                        "Tiempo (min)",
                        "Tasa de Éxito (%)",
                    ]
                ]
                for s in student_rows:
                    resumen_data.append(
                        [
                            s["nombre"],
                            f"{s['promedio']:.2f}",
                            f"{s['mejor']:.2f}",
                            str(s["completados"]),
                            str(s["intentos"]),
                            f"{s['tiempo_min']:.2f}",
                            f"{s['tasa']:.1f}%",
                        ]
                    )

                resumen_table = Table(
                    resumen_data,
                    colWidths=[
                        1.8 * inch,
                        0.8 * inch,
                        0.9 * inch,
                        0.9 * inch,
                        0.8 * inch,
                        0.9 * inch,
                        1.0 * inch,
                    ],
                )
                resumen_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C5AA0")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 9),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F5FB")]),
                        ]
                    )
                )
                elements.append(resumen_table)
                elements.append(Spacer(1, 0.3 * inch))

                # --- Desglose por Estudiante (tabla detallada) ---
                elements.append(Paragraph("Desempeño por Estudiante", heading_style))

                detail_data = [
                    [
                        "Estudiante",
                        "Matrícula",
                        "Intentos",
                        "Completados",
                        "Promedio",
                        "Tiempo",
                        "Tasa Éxito (%)",
                    ]
                ]
                for s in student_rows:
                    tiempo_seg_total = int(s["tiempo_min"] * 60)
                    detail_data.append(
                        [
                            s["nombre"],
                            s["matricula"],
                            str(s["intentos"]),
                            str(s["completados"]),
                            f"{s['promedio']:.2f}",
                            f"{tiempo_seg_total // 60}m {tiempo_seg_total % 60}s",
                            f"{s['tasa']:.1f}%",
                        ]
                    )

                table = Table(
                    detail_data,
                    colWidths=[
                        1.8 * inch,
                        1 * inch,
                        0.8 * inch,
                        1 * inch,
                        0.8 * inch,
                        1 * inch,
                        1 * inch,
                    ],
                )
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C5AA0")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 9),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F5FB")]),
                        ]
                    )
                )
                elements.append(table)

            elements.append(Spacer(1, 0.2 * inch))

            # Pie de página
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(
                Paragraph(
                    f"<font size=8>Reporte generado automáticamente el "
                    f"{datetime.now().strftime('%Y-%m-%d %H:%M')}</font>",
                    styles["Normal"],
                )
            )

            # Construir PDF
            doc.build(elements)

        except Exception:
            logger.exception("Error generando PDF del salón %s", salon_id)
            raise

        return pdf_buffer.getvalue()

    @staticmethod
    async def generate_student_pdf_report(db: AsyncSession, alumno_id: str, salon_id: str) -> bytes:
        """
        Genera un reporte PDF de expediente del estudiante.
        Incluye detalles de intentos, parámetros físicos y tiempo de resolución.
        """
        # Obtener estudiante e interacciones
        result = await db.execute(
            select(Alumno).where(Alumno.idalumno == alumno_id).options(selectinload(Alumno.usuario))
        )
        alumno = result.scalar_one_or_none()
        if not alumno:
            raise ValueError(f"Estudiante {alumno_id} no encontrado")

        # Obtener interacciones del salón
        result = await db.execute(
            select(InteraccionEscenario)
            .where(
                InteraccionEscenario.idalumno == alumno_id,
                InteraccionEscenario.idescenario.in_(
                    select(Escenario.idescenario).where(Escenario.idsalon == salon_id)
                ),
            )
            .options(selectinload(InteraccionEscenario.escenario))
        )
        interacciones = result.scalars().all()

        usuario = alumno.usuario

        # Crear documento PDF
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=letter,
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=22,
            textColor=colors.HexColor("#1F4788"),
            spaceAfter=12,
            alignment=1,
        )

        elements = []

        # Título
        elements.append(Paragraph("PARABOLIC LAB", title_style))
        elements.append(Paragraph("Expediente de Estudiante", styles["Heading2"]))
        elements.append(Spacer(1, 0.2 * inch))

        # Información del estudiante
        nombre_completo = " ".join(
            filter(None, [usuario.nombre, usuario.apellidopaterno, usuario.apellidomaterno])
        )
        student_info = [
            ["Nombre:", nombre_completo or "Sin nombre"],
            ["Matrícula:", alumno.matricula],
            ["Email:", usuario.email or "N/A"],
            ["Fecha Reporte:", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ]
        student_table = Table(student_info, colWidths=[2 * inch, 4 * inch])
        student_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8F0F8")),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ]
            )
        )
        elements.append(student_table)
        elements.append(Spacer(1, 0.3 * inch))

        if not interacciones:
            elements.append(
                Paragraph(
                    "Sin datos registrados - Este estudiante aún no ha participado en escenarios.", styles["Normal"]
                )
            )
        else:
            # --- Resumen General de Desempeño (upfront) ---
            heading_style = ParagraphStyle(
                "CustomHeading",
                parent=styles["Heading2"],
                fontSize=14,
                textColor=colors.HexColor("#2C5AA0"),
                spaceAfter=10,
                spaceBefore=10,
            )
            elements.append(Paragraph("Resumen General de Desempeño", heading_style))

            total_intentos_resumen = sum(int(i.intentosrealizados or 0) for i in interacciones)
            completadas_resumen = sum(1 for i in interacciones if i.completado)
            puntuacion_total_resumen = sum(float(i.puntuacion or 0) for i in interacciones)
            num_interacciones = len(interacciones)
            promedio_resumen = puntuacion_total_resumen / num_interacciones if num_interacciones > 0 else 0
            mejor_resumen = max((float(i.puntuacion or 0) for i in interacciones), default=0)
            tiempo_total_resumen = sum(int(i.tiempototal or 0) for i in interacciones)

            tasa_exito_resumen = (
                completadas_resumen / num_interacciones * 100 if num_interacciones > 0 else 0
            )
            resumen_data = [
                ["Promedio de Puntuación:", f"{promedio_resumen:.1f}"],
                ["Mejor Puntuación:", f"{mejor_resumen:.1f}"],
                ["Escenarios Completados:", f"{completadas_resumen} / {num_interacciones}"],
                ["Intentos Totales:", str(total_intentos_resumen)],
                ["Tiempo Total:", f"{tiempo_total_resumen} seg ({tiempo_total_resumen // 60} min)"],
                ["Tasa de Éxito:", f"{tasa_exito_resumen:.1f}%"],
            ]
            resumen_table = Table(resumen_data, colWidths=[2.5 * inch, 3.5 * inch])
            resumen_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8F0F8")),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ]
                )
            )
            elements.append(resumen_table)
            elements.append(Spacer(1, 0.3 * inch))

            # --- Detalle de Interacciones ---
            elements.append(Paragraph("Detalle de Interacciones", heading_style))

            # Tabla de intentos detallada
            table_data = [
                ["Escenario", "Nivel", "v₀ (m/s)", "θ (°)", "Intentos", "Completado", "Puntos", "Tiempo (s)", "Fecha"]
            ]

            for interaccion in interacciones:
                escenario = interaccion.escenario
                datos = interaccion.datosinteraccion or {}

                v0 = str(datos.get("velocidadInicial", "N/A"))
                angulo = str(datos.get("angulo", "N/A"))

                table_data.append(
                    [
                        escenario.nombre[:30],
                        escenario.niveldificultad[:3],
                        v0,
                        angulo,
                        str(interaccion.intentosrealizados or 0),
                        "✓" if interaccion.completado else "✗",
                        f"{float(interaccion.puntuacion or 0):.1f}",
                        f"{int(interaccion.tiempototal or 0)}",
                        interaccion.fechafin.strftime("%Y-%m-%d %H:%M") if interaccion.fechafin else "Sin registro",
                    ]
                )

            table = Table(
                table_data,
                colWidths=[
                    1.5 * inch,
                    0.7 * inch,
                    0.7 * inch,
                    0.6 * inch,
                    0.7 * inch,
                    0.75 * inch,
                    0.6 * inch,
                    0.75 * inch,
                    0.75 * inch,
                ],
            )
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C5AA0")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 1), 8),
                        ("FONTSIZE", (0, 2), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F5FB")]),
                    ]
                )
            )
            elements.append(table)

            # The summary stats are now shown at the top of the report.
            # No duplicate stats section at the bottom.

        # Pie de página
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(
            Paragraph(
                f"<font size=8>Reporte generado automáticamente el "
                f"{datetime.now().strftime('%Y-%m-%d %H:%M')}</font>",
                styles["Normal"],
            )
        )

        doc.build(elements)
        return pdf_buffer.getvalue()

    @staticmethod
    async def generate_interaccion_pdf_report(reporte: ReporteInteraccionRead) -> bytes:
        """Genera PDF del reporte de un intento completado."""

        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=letter,
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=22,
            textColor=colors.HexColor("#1F4788"),
            spaceAfter=12,
            alignment=1,
        )
        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=13,
            textColor=colors.HexColor("#2C5AA0"),
            spaceAfter=8,
            spaceBefore=10,
        )
        elements = []

        def fmt_val(v: float | None, decimales: int = 1) -> str:
            return f"{v:.{decimales}f}" if v is not None else "—"

        def fmt_dt(dt: datetime | None) -> str:
            return dt.strftime("%Y-%m-%d %H:%M") if dt else "—"

        def fmt_tiempo(seg: int | None) -> str:
            seg = seg or 0
            return f"{seg // 60}:{seg % 60:02d}"

        # --- Cabecera (igual que el reporte en pantalla) ---
        elements.append(Paragraph("PARABOLIC LAB", title_style))
        elements.append(Paragraph(reporte.nombre_escenario, styles["Heading2"]))
        elements.append(
            Paragraph(
                f"<font size=9 color='#6B7280'>Nivel: {reporte.niveldificultad} · "
                f"Finalizado: {fmt_dt(reporte.fechafin)}</font>",
                styles["Normal"],
            )
        )
        elements.append(Spacer(1, 0.2 * inch))

        # --- Resumen numérico (tarjetas) ---
        aciertos = sum(1 for d in reporte.disparos if d.acierto)
        card_label = ParagraphStyle(
            "CardLabel", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#6B7280")
        )
        card_value = ParagraphStyle(
            "CardValue", parent=styles["Normal"], fontSize=16, leading=18,
            textColor=colors.HexColor("#1F2937"), fontName="Helvetica-Bold",
        )

        def _card(label: str, value: str) -> list:
            return [Paragraph(label.upper(), card_label), Paragraph(value, card_value)]

        resumen_data = [[
            _card("Puntuación", fmt_val(reporte.puntuacion, 1)),
            _card("Tiempo", fmt_tiempo(reporte.tiempototal)),
            _card("Intentos", str(reporte.intentosrealizados or 0)),
            _card("Aciertos", f"{aciertos}/{len(reporte.disparos)}"),
        ]]
        resumen_table = Table(resumen_data, colWidths=[1.7 * inch] * 4)
        resumen_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F0F5FB")),
                ("BOX", (0, 0), (0, -1), 0.5, colors.HexColor("#D1D9E6")),
                ("BOX", (1, 0), (1, -1), 0.5, colors.HexColor("#D1D9E6")),
                ("BOX", (2, 0), (2, -1), 0.5, colors.HexColor("#D1D9E6")),
                ("BOX", (3, 0), (3, -1), 0.5, colors.HexColor("#D1D9E6")),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ])
        )
        elements.append(resumen_table)
        elements.append(Spacer(1, 0.25 * inch))

        # --- Análisis: comparación alumno vs solución ---
        elements.append(Paragraph("Análisis — Tus respuestas vs la solución correcta", heading_style))
        analisis = reporte.analisis
        analisis_data = [
            ["Variable", "Tu respuesta", "Valor correcto"],
            ["Ángulo (°)", fmt_val(analisis.angulo_alumno), fmt_val(analisis.angulo_correcto)],
            ["Velocidad inicial (m/s)", fmt_val(analisis.velocidad_alumno), fmt_val(analisis.velocidad_correcta)],
            ["Alcance (m)", fmt_val(analisis.alcance_alumno), fmt_val(analisis.alcance_correcto)],
            ["Altura máxima (m)", fmt_val(analisis.altura_maxima_alumno), "—"],
            ["Tiempo de vuelo (s)", fmt_val(analisis.tiempo_vuelo_alumno), "—"],
        ]
        analisis_table = Table(analisis_data, colWidths=[2.5 * inch, 1.9 * inch, 1.9 * inch])
        analisis_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C5AA0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F5FB")]),
            ])
        )
        elements.append(analisis_table)
        elements.append(Spacer(1, 0.15 * inch))

        # --- Procedimiento y notas del alumno ---
        def _bloque_texto(titulo: str, texto: str, font_name: str) -> list:
            """Sección con título de encabezado y texto en una caja con fondo.

            El fondo se dibuja con una Table de una celda: usar backColor +
            borderPadding en un Paragraph pinta el fondo fuera de su marco y
            encima del título anterior, cortándolo visualmente.
            """
            body_style = ParagraphStyle(
                f"Bloque{titulo}", parent=styles["Normal"], fontName=font_name,
                fontSize=9, leading=12,
            )
            # Conservar saltos de línea del texto del alumno.
            html = texto.replace("&", "&amp;").replace("<", "&lt;").replace("\n", "<br/>")
            caja = Table([[Paragraph(html, body_style)]], colWidths=[7.5 * inch])
            caja.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F0F5FB")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D9E6")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ])
            )
            return [Paragraph(titulo, heading_style), caja, Spacer(1, 0.15 * inch)]

        if analisis.procedimiento:
            elements.extend(_bloque_texto("Procedimiento", analisis.procedimiento, "Courier"))

        if analisis.notas:
            elements.extend(_bloque_texto("Notas", analisis.notas, "Helvetica"))

        elements.append(Spacer(1, 0.1 * inch))

        # --- Disparos ---
        if reporte.disparos:
            elements.append(Paragraph(f"Historial de Disparos ({len(reporte.disparos)})", heading_style))
            disparos_data: list[list[str]] = [
                ["#", "Ángulo (°)", "Velocidad (m/s)", "Alt. cañón (m)", "Distancia (m)", "Resultado", "Puntos"]
            ]
            for d in reporte.disparos:
                disparos_data.append([
                    str(d.n),
                    fmt_val(d.angulo, 1),
                    fmt_val(d.velocidad, 1),
                    fmt_val(d.altura_canon, 1),
                    fmt_val(d.distancia, 1),
                    "✓ Acierto" if d.acierto else "✗ Falló",
                    str(d.puntos),
                ])
            disparos_table = Table(
                disparos_data,
                colWidths=[0.4 * inch, 0.9 * inch, 0.9 * inch, 0.9 * inch, 1.0 * inch, 0.9 * inch, 0.7 * inch],
            )
            disparos_table.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C5AA0")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F5FB")]),
                ])
            )
            elements.append(disparos_table)
            elements.append(Spacer(1, 0.25 * inch))

        # --- Comparativa con intentos anteriores ---
        if reporte.comparativa:
            elements.append(Paragraph("Comparativa — Intentos anteriores en este escenario", heading_style))
            comp_data = [["Fecha inicio", "Puntuación", "Intentos", "Estado"]]
            for c in reporte.comparativa:
                comp_data.append([
                    fmt_dt(c.fechainicio),
                    fmt_val(c.puntuacion, 1),
                    str(c.intentosrealizados or 0),
                    "Completado" if c.completado else "En progreso",
                ])
            comp_table = Table(comp_data, colWidths=[2 * inch, 1.5 * inch, 1 * inch, 1.5 * inch])
            comp_table.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C5AA0")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F5FB")]),
                ])
            )
            elements.append(comp_table)

        elements.append(Spacer(1, 0.3 * inch))
        elements.append(
            Paragraph(
                f"<font size=8>Reporte generado el {datetime.now().strftime('%Y-%m-%d %H:%M')}</font>",
                styles["Normal"],
            )
        )
        doc.build(elements)
        return pdf_buffer.getvalue()

    @staticmethod
    async def generate_interaccion_xlsx_report(reporte: ReporteInteraccionRead) -> bytes:
        """Genera Excel del reporte de un intento completado."""
        wb = openpyxl.Workbook()

        HEADER_FILL = PatternFill("solid", fgColor="2C5AA0")
        HEADER_FONT = Font(bold=True, color="FFFFFF")
        LABEL_FILL = PatternFill("solid", fgColor="E8F0F8")
        BOLD = Font(bold=True)
        CENTER = Alignment(horizontal="center")

        def fmt_val(v: float | None, decimales: int = 1) -> str:
            return f"{v:.{decimales}f}" if v is not None else "N/A"

        def fmt_dt(dt: datetime | None) -> str:
            return dt.strftime("%Y-%m-%d %H:%M") if dt else "—"

        # --- Hoja 1: Datos generales ---
        ws = wb.active
        ws.title = "Reporte"
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 20

        def write_section_header(row: int, title: str) -> int:
            cell = ws.cell(row=row, column=1, value=title)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            return row + 1

        def write_row(row: int, label: str, value: str) -> int:
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = BOLD
            lc.fill = LABEL_FILL
            ws.cell(row=row, column=2, value=value)
            return row + 1

        r = 1
        r = write_section_header(r, "PARABOLIC LAB — Reporte de Intento")
        r += 1
        r = write_section_header(r, "Datos del intento")
        r = write_row(r, "Escenario", reporte.nombre_escenario)
        r = write_row(r, "Nivel de dificultad", reporte.niveldificultad)
        r = write_row(r, "Fecha inicio", fmt_dt(reporte.fechainicio))
        r = write_row(r, "Fecha fin", fmt_dt(reporte.fechafin))
        r = write_row(r, "Tiempo total (s)", str(reporte.tiempototal or 0))
        r = write_row(r, "Intentos realizados", str(reporte.intentosrealizados or 0))
        r = write_row(r, "Puntuación", fmt_val(reporte.puntuacion))
        r += 1

        # Análisis
        r = write_section_header(r, "Análisis — Alumno vs Solución Correcta")
        analisis = reporte.analisis
        headers = ["Variable", "Valor alumno", "Valor correcto"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=col, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER
        ws.column_dimensions["C"].width = 18
        r += 1
        analisis_rows = [
            ("Ángulo (°)", fmt_val(analisis.angulo_alumno), fmt_val(analisis.angulo_correcto)),
            ("Velocidad inicial (m/s)", fmt_val(analisis.velocidad_alumno), fmt_val(analisis.velocidad_correcta)),
            ("Alcance (m)", fmt_val(analisis.alcance_alumno), fmt_val(analisis.alcance_correcto)),
            ("Altura máxima (m)", fmt_val(analisis.altura_maxima_alumno), "—"),
            ("Tiempo de vuelo (s)", fmt_val(analisis.tiempo_vuelo_alumno), "—"),
        ]
        for label, alumno_val, correcto_val in analisis_rows:
            lc = ws.cell(row=r, column=1, value=label)
            lc.font = BOLD
            ws.cell(row=r, column=2, value=alumno_val).alignment = CENTER
            ws.cell(row=r, column=3, value=correcto_val).alignment = CENTER
            r += 1
        r += 1

        # Procedimiento y notas del alumno
        if analisis.procedimiento:
            r = write_section_header(r, "Procedimiento")
            pc = ws.cell(row=r, column=1, value=analisis.procedimiento)
            pc.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            r += 2
        if analisis.notas:
            r = write_section_header(r, "Notas")
            nc = ws.cell(row=r, column=1, value=analisis.notas)
            nc.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            r += 2

        # --- Hoja 2: Disparos ---
        if reporte.disparos:
            ws2 = wb.create_sheet("Disparos")
            ws2.column_dimensions["A"].width = 6
            ws2.column_dimensions["B"].width = 14
            ws2.column_dimensions["C"].width = 18
            ws2.column_dimensions["D"].width = 16
            ws2.column_dimensions["E"].width = 16
            ws2.column_dimensions["F"].width = 12
            ws2.column_dimensions["G"].width = 10
            disp_headers = [
                "#", "Ángulo (°)", "Velocidad (m/s)", "Alt. cañón (m)", "Distancia (m)", "Resultado", "Puntos"
            ]
            for col, h in enumerate(disp_headers, 1):
                c = ws2.cell(row=1, column=col, value=h)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = CENTER
            for i, d in enumerate(reporte.disparos, 2):
                ws2.cell(row=i, column=1, value=d.n)
                ws2.cell(row=i, column=2, value=fmt_val(d.angulo))
                ws2.cell(row=i, column=3, value=fmt_val(d.velocidad))
                ws2.cell(row=i, column=4, value=fmt_val(d.altura_canon))
                ws2.cell(row=i, column=5, value=fmt_val(d.distancia))
                ws2.cell(row=i, column=6, value="Acierto" if d.acierto else "Falló")
                ws2.cell(row=i, column=7, value=d.puntos)

        # --- Hoja 3: Comparativa ---
        if reporte.comparativa:
            ws3 = wb.create_sheet("Comparativa")
            ws3.column_dimensions["A"].width = 20
            ws3.column_dimensions["B"].width = 14
            ws3.column_dimensions["C"].width = 12
            ws3.column_dimensions["D"].width = 14
            comp_headers = ["Fecha inicio", "Puntuación", "Intentos", "Estado"]
            for col, h in enumerate(comp_headers, 1):
                c = ws3.cell(row=1, column=col, value=h)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = CENTER
            for i, c_obj in enumerate(reporte.comparativa, 2):
                ws3.cell(row=i, column=1, value=fmt_dt(c_obj.fechainicio))
                ws3.cell(row=i, column=2, value=fmt_val(c_obj.puntuacion))
                ws3.cell(row=i, column=3, value=c_obj.intentosrealizados or 0)
                ws3.cell(row=i, column=4, value="Completado" if c_obj.completado else "En progreso")

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
